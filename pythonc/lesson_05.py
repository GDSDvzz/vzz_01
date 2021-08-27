"""
============================
Project: python_class
Author:柠檬班-Tricy
Time:2021/8/19 17:54
E-mail:3247119728@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
'''
1：a=[1,2,'6','summer'] 请用成员运算符判断 i是否在这个列表里面 -- if
2：dict_1={"class_id":45,'num':20} 请判断班级上课人数是否大于5
注：num表示课堂人数。如果大于5，输出人数。
3、list1 = ['Rella', '纯纯', '二丫', '小梨', '东东', '熊熊']
列表当中的每一个值包含：姓名、性别、年龄、城市。以字典的形式表达。并且把字典都存在新的 list中，最后打印最终的列表。--list2=[]
方法1： 手动扩充--字典--存放在列表里面；{} --简单
方法2： 自动--dict（）--不强制-- for循环 ，list2.append() --- 难度
4、for循环遍历其他的数据类型 --字典，列表 元组
'''
# a=[1,2,'6','summer']
# if "i" in a:
#     print("i是这个列表的成员！")
# else:
#     print("i不是列表的成员")
#
# dict_1={"class_id":45,'num':20}
# num = dict_1.get("num")
# if num > 5:
#     print("这个班级的人数是：{}".format(num))
# else:
#     print("班级人数不足5人！")

# 不唯一的
# dict1 = {"name":"Rella","gender":"male","age":18,"city":"深圳"}
# dict2 = {"name":"纯纯","gender":"male","age":18,"city":"深圳"}
# dict3 = {"name":"二丫","gender":"male","age":18,"city":"深圳"}
# dict4 = {"name":"小梨","gender":"male","age":18,"city":"深圳"}
# dict5 = {"name":"东东","gender":"male","age":18,"city":"深圳"}
# dict6 = {"name":"熊熊","gender":"male","age":18,"city":"深圳"}
# list2 = [dict1,dict2,dict3,dict4,dict5,dict6]
# print(list2)
#
# list1 = ['Rella', '纯纯', '二丫', '小梨', '东东', '熊熊']
# list2 = [] #空列表
# for i in range(len(list1)):
#     dict1 = dict(name=list1[i],age=18,gender="male",city="北京")
#     list2.append(dict1)
# print(list2)

"""
接口自动化的测试思路：
1、用例excel  --done
2、自动去excel读取数据 --done== read_date()
3、发送接口测试  ---done 
4、测试结果 vs  执行结果 --断言  --通过  不通过
5、结果回写测试用例 ---done 

读取excel表格里的测试数据 --- openpyxl ==第三方库
1、pip install openpyxl --pycharm安装
2、导入 --import openpyxl
excel 三大部分：
1）工作簿对象
2) 表单 --sheet
3） 单元格 -cell
4） 获取内容-- .value
5) 写入== 获取的值重新赋值  ==保存生效 （关闭文件）
cell = sheet.cell(row=1,column=1) # 单元格
print(cell.value)  # 单元格的内容
cell.value = "测试用例"  # 写入单元格的内容
wb.save("testcase_api_wuye.xlsx")  # 保存文件，另存为===  注意关闭文件，否则报错！
print(cell.value)

封装函数步骤：
1、功能写出来--def
2、变化的数据  -- 参数化
3、给调用的人使用数据？ -- 返回值
"""
from openpyxl import load_workbook  # 导入部分
import requests
# 读取函数
def read_data(filename,sheetname):
    wb = load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname] # 表单
    max_row = sheet.max_row  #获取最大行数
    cases = [] # 空列表--所有测试用例
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expected_result = sheet.cell(row=i,column=7).value)
        cases.append(case)
    return cases

# 发送接口请求函数
def api_request(url_wuye,test_data):
    head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  # 定义请求头部
    response = requests.post(url=url_wuye,json=test_data,headers=head)  # 发送接口请求的 --响应消息
    return  response.json()   # 接口的响应结果json

# 回写结果函数
def write_date(filename,sheetname,row,column,final_result):
    wb = load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname] # 表单
    sheet.cell(row,column).value = final_result  # 写入
    wb.save(filename)

# 执行测试
'''
{'case_id': 1, 
'url': 'http://47.115.15.198:7001/smarthome/user/register', 
'data': 
'{"phone":"17750542236","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬小可爱1","verificationCode":"611203"}',
'expected_result': '{"code":"0","msg":"操作成功"}'}
注意：数据从excel 表格读取-- 文本 =字符串
eval（）===  脱引号 --- 字符串包裹的python表达式--类型转化 
'''
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get("case_id")
        url = case["url"]
        data = case["data"]  # 字符串
        data = eval(data)  # 字符串强制转化为 字典
        expected_result = case["expected_result"]
        expected_result = eval(expected_result) #  字符串强制转化为 字典
        expect_msg = expected_result["msg"]
        real_result = api_request(url_wuye=url,test_data=data)  # 调用接口请求函数
        real_msg = real_result["msg"]
        print("预期结果是：{}".format(expect_msg))
        print("执行结果是：{}".format(real_msg))
        if expect_msg == real_msg:
            print("第{}条测试用例通过！".format(case_id))
            final_result = "Passed"
        else:
            print("第{}条测试用例不通过！".format(case_id))
            final_result = "Failed"
        print("-*" * 20)
        write_date(filename,sheetname,case_id+1,8,final_result=final_result)  # 写入结果函数

execute_fun("testcase_api_wuye.xlsx","login")

"""
selenium : 第三方库/工具 == pip install selenium
"""







